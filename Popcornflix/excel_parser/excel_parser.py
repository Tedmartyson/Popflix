from openpyxl import load_workbook

def excel_parse(excel_file):
    excel_houses = []

    wb = load_workbook(excel_file)
    ws = wb['Лист1']
    for row in ws.iter_rows(min_row=2, max_col=8):
        if row[0].internal_value is None:
            break
        address, _, rooms, area, floor, floor_number, price, description = [r.internal_value for r in row]
    
        try:
            rooms = int(rooms)
        except ValueError:
            rooms = 0
        
        excel_houses.append({
            'address': address,
            'rooms': rooms,
            'area': area,
            'floor': floor,
            'floor_number': floor_number,
            'price': price,
            'description': description
        })

    return excel_houses

    # h = House(address=address, rooms=rooms, area=area, floor=floor, floor_number=floor_number, price=price, description=description)
    # h.save()
    